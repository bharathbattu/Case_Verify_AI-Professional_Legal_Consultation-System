"""
Excel Export functionality for Case-Verify AI
"""
import io
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from database.models import Case, Analysis, User

class ExcelExporter:
    """Excel export functionality for cases and analysis data"""
    
    def __init__(self):
        self.setup_styles()
    
    def setup_styles(self):
        """Setup Excel styles"""
        # Header style
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        
        # Subheader style
        self.subheader_font = Font(name='Arial', size=11, bold=True, color='2F5597')
        self.subheader_fill = PatternFill(start_color='E6F1FF', end_color='E6F1FF', fill_type='solid')
        
        # Data style
        self.data_font = Font(name='Arial', size=10)
        
        # Border
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alignment
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    def export_cases_to_excel(self, cases: List[Case], user: User = None) -> io.BytesIO:
        """Export cases to Excel format"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, cases, user)
        self._create_cases_detail_sheet(wb, cases)
        self._create_statistics_sheet(wb, cases)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_summary_sheet(self, wb: Workbook, cases: List[Case], user: User = None):
        """Create summary sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Case-Verify AI - Export Summary"
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='2F5597')
        ws.merge_cells('A1:F1')
        
        # Export information
        row = 3
        ws[f'A{row}'] = "Export Date:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        row += 1
        ws[f'A{row}'] = "Total Cases:"
        ws[f'B{row}'] = len(cases)
        
        if user:
            row += 1
            ws[f'A{row}'] = "User:"
            ws[f'B{row}'] = user.full_name or user.username
        
        # Statistics
        row += 2
        ws[f'A{row}'] = "Case Statistics"
        ws[f'A{row}'].font = self.subheader_font
        
        # Count by status
        status_counts = {}
        priority_counts = {}
        type_counts = {}
        
        for case in cases:
            status_counts[case.status] = status_counts.get(case.status, 0) + 1
            priority_counts[case.priority] = priority_counts.get(case.priority, 0) + 1
            if case.case_type:
                type_counts[case.case_type] = type_counts.get(case.case_type, 0) + 1
        
        # Status breakdown
        row += 2
        ws[f'A{row}'] = "Status Breakdown:"
        ws[f'A{row}'].font = Font(bold=True)
        
        for status, count in status_counts.items():
            row += 1
            ws[f'A{row}'] = f"  {status.title()}:"
            ws[f'B{row}'] = count
        
        # Priority breakdown
        row += 2
        ws[f'A{row}'] = "Priority Breakdown:"
        ws[f'A{row}'].font = Font(bold=True)
        
        for priority, count in priority_counts.items():
            row += 1
            ws[f'A{row}'] = f"  {priority.title()}:"
            ws[f'B{row}'] = count
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_cases_detail_sheet(self, wb: Workbook, cases: List[Case]):
        """Create detailed cases sheet"""
        ws = wb.create_sheet("Case Details")
        
        # Headers
        headers = [
            'Case ID', 'Title', 'Status', 'Priority', 'Case Type', 'Case Category',
            'PIN Code', 'Limitation Period', 'Days Remaining', 'Court Suggestion',
            'Confidence Score', 'Created Date', 'Updated Date', 'Facts', 'Relief Sought'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Data rows
        for row, case in enumerate(cases, 2):
            data = [
                case.id,
                case.title,
                case.status,
                case.priority,
                case.case_type or '',
                case.case_category or '',
                case.pin_code,
                case.limitation_period or '',
                case.days_remaining if case.days_remaining is not None else '',
                case.court_suggestion or '',
                case.confidence_score if case.confidence_score is not None else '',
                case.created_at.strftime('%Y-%m-%d') if case.created_at else '',
                case.updated_at.strftime('%Y-%m-%d') if case.updated_at else '',
                case.facts,
                case.relief_sought
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.data_font
                cell.border = self.thin_border
                
                # Special formatting for text columns
                if col in [14, 15]:  # Facts and Relief columns
                    cell.alignment = self.left_alignment
                else:
                    cell.alignment = self.center_alignment
                
                # Color coding for status
                if col == 3:  # Status column
                    if value == 'draft':
                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    elif value == 'analyzed':
                        cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                    elif value == 'filed':
                        cell.fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
                    elif value == 'closed':
                        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                
                # Color coding for priority
                if col == 4:  # Priority column
                    if value == 'urgent':
                        cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                    elif value == 'high':
                        cell.fill = PatternFill(start_color='FFEECC', end_color='FFEECC', fill_type='solid')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set specific widths for certain columns
            if column_letter in ['A']:  # Case ID
                adjusted_width = 10
            elif column_letter in ['B']:  # Title
                adjusted_width = 25
            elif column_letter in ['N', 'O']:  # Facts, Relief
                adjusted_width = 40
            else:
                adjusted_width = min(max_length + 2, 20)
            
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    def _create_statistics_sheet(self, wb: Workbook, cases: List[Case]):
        """Create statistics sheet with charts"""
        ws = wb.create_sheet("Statistics")
        
        # Title
        ws['A1'] = "Case Statistics Dashboard"
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='2F5597')
        ws.merge_cells('A1:F1')
        
        # Prepare data
        status_counts = {}
        priority_counts = {}
        monthly_counts = {}
        
        for case in cases:
            # Status counts
            status_counts[case.status] = status_counts.get(case.status, 0) + 1
            
            # Priority counts
            priority_counts[case.priority] = priority_counts.get(case.priority, 0) + 1
            
            # Monthly counts
            if case.created_at:
                month_key = case.created_at.strftime('%Y-%m')
                monthly_counts[month_key] = monthly_counts.get(month_key, 0) + 1
        
        # Status statistics table
        row = 3
        ws[f'A{row}'] = "Status Distribution"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        
        row += 1
        ws[f'A{row}'] = "Status"
        ws[f'B{row}'] = "Count"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        status_data_start = row + 1
        for status, count in status_counts.items():
            row += 1
            ws[f'A{row}'] = status.title()
            ws[f'B{row}'] = count
        
        # Priority statistics table
        row += 3
        ws[f'A{row}'] = "Priority Distribution"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        
        row += 1
        ws[f'A{row}'] = "Priority"
        ws[f'B{row}'] = "Count"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        priority_data_start = row + 1
        for priority, count in priority_counts.items():
            row += 1
            ws[f'A{row}'] = priority.title()
            ws[f'B{row}'] = count
        
        # Create charts if we have data
        if status_counts:
            self._create_pie_chart(ws, status_data_start, len(status_counts), 'D3', 'Status Distribution')
        
        if priority_counts:
            self._create_bar_chart(ws, priority_data_start, len(priority_counts), 'D15', 'Priority Distribution')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_pie_chart(self, ws, data_start_row: int, data_count: int, position: str, title: str):
        """Create pie chart"""
        try:
            chart = PieChart()
            chart.title = title
            chart.height = 10
            chart.width = 15
            
            # Data references
            labels = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_start_row + data_count - 1)
            data = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_start_row + data_count - 1)
            
            chart.add_data(data)
            chart.set_categories(labels)
            
            ws.add_chart(chart, position)
        except Exception as e:
            # If chart creation fails, just skip it
            pass
    
    def _create_bar_chart(self, ws, data_start_row: int, data_count: int, position: str, title: str):
        """Create bar chart"""
        try:
            chart = BarChart()
            chart.title = title
            chart.height = 10
            chart.width = 15
            
            # Data references
            labels = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_start_row + data_count - 1)
            data = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_start_row + data_count - 1)
            
            chart.add_data(data)
            chart.set_categories(labels)
            
            ws.add_chart(chart, position)
        except Exception as e:
            # If chart creation fails, just skip it
            pass
    
    def export_single_case_to_excel(self, case: Case, analysis: Analysis = None) -> io.BytesIO:
        """Export single case with analysis to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Case Analysis"
        
        # Title
        ws['A1'] = f"Case Analysis Report: {case.title}"
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='2F5597')
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Case Information
        ws[f'A{row}'] = "Case Information"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        case_info = [
            ('Case ID:', case.id),
            ('Title:', case.title),
            ('Status:', case.status),
            ('Priority:', case.priority),
            ('Case Type:', case.case_type or 'Not specified'),
            ('PIN Code:', case.pin_code),
            ('Created Date:', case.created_at.strftime('%Y-%m-%d') if case.created_at else ''),
            ('Limitation Period:', case.limitation_period or 'Not determined'),
            ('Days Remaining:', case.days_remaining if case.days_remaining is not None else 'Not calculated'),
            ('Court Suggestion:', case.court_suggestion or 'Not determined'),
        ]
        
        for label, value in case_info:
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
        
        # Facts
        row += 2
        ws[f'A{row}'] = "Facts of the Case"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = case.facts
        ws[f'A{row}'].alignment = self.left_alignment
        ws.merge_cells(f'A{row}:D{row}')
        
        # Relief Sought
        row += 2
        ws[f'A{row}'] = "Relief Sought"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = case.relief_sought
        ws[f'A{row}'].alignment = self.left_alignment
        ws.merge_cells(f'A{row}:D{row}')
        
        # Analysis Information (if available)
        if analysis:
            row += 2
            ws[f'A{row}'] = "AI Analysis Results"
            ws[f'A{row}'].font = self.subheader_font
            ws[f'A{row}'].fill = self.subheader_fill
            ws.merge_cells(f'A{row}:D{row}')
            
            analysis_info = [
                ('Analysis Date:', analysis.created_at.strftime('%Y-%m-%d %H:%M:%S') if analysis.created_at else ''),
                ('Confidence Score:', f"{analysis.confidence_score:.1f}/10" if analysis.confidence_score else 'N/A'),
                ('Processing Time:', f"{analysis.processing_time:.2f} seconds" if analysis.processing_time else 'N/A'),
                ('AI Model Used:', analysis.ai_model_used or 'Not specified'),
            ]
            
            for label, value in analysis_info:
                row += 1
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
            
            # Legal Reasoning
            if analysis.legal_reasoning:
                row += 2
                ws[f'A{row}'] = "Legal Reasoning"
                ws[f'A{row}'].font = self.subheader_font
                ws[f'A{row}'].fill = self.subheader_fill
                ws.merge_cells(f'A{row}:D{row}')
                
                row += 1
                ws[f'A{row}'] = analysis.legal_reasoning
                ws[f'A{row}'].alignment = self.left_alignment
                ws.merge_cells(f'A{row}:D{row}')
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def create_dataframe_from_cases(self, cases: List[Case]) -> pd.DataFrame:
        """Create pandas DataFrame from cases for analysis"""
        data = []
        
        for case in cases:
            data.append({
                'Case ID': case.id,
                'Title': case.title,
                'Status': case.status,
                'Priority': case.priority,
                'Case Type': case.case_type or '',
                'Case Category': case.case_category or '',
                'PIN Code': case.pin_code,
                'Limitation Period': case.limitation_period or '',
                'Days Remaining': case.days_remaining,
                'Court Suggestion': case.court_suggestion or '',
                'Confidence Score': case.confidence_score,
                'Created Date': case.created_at.strftime('%Y-%m-%d') if case.created_at else '',
                'Updated Date': case.updated_at.strftime('%Y-%m-%d') if case.updated_at else '',
                'Facts Length': len(case.facts) if case.facts else 0,
                'Relief Length': len(case.relief_sought) if case.relief_sought else 0,
                'Tags Count': len(case.tags) if case.tags else 0
            })
        
        return pd.DataFrame(data)
